from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework import status
from .serializers import *
from .models import *
from Auth_LDAP.models import *
from .tasks import create_task
from openpyxl import load_workbook
from rest_framework.views import APIView
import json
from django.db.models import Max
import os
import io
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection
from mui_table_settings import filter_queryset, CustomPagination

def copy_row(sheet, row_to_copy):
    # Вставляем новую строку выше текущей
    sheet.insert_rows(row_to_copy)

    # Копируем данные и стили из строки row_to_copy + 1 (так как мы вставили новую строку)
    for col in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(row=row_to_copy + 1, column=col)
        target_cell = sheet.cell(row=row_to_copy, column=col)

        # Копируем стили
        if source_cell.has_style:
            target_cell.font = Font(name=source_cell.font.name,
                                     size=source_cell.font.size,
                                     bold=source_cell.font.bold,
                                     italic=source_cell.font.italic,
                                     vertAlign=source_cell.font.vertAlign,
                                     underline=source_cell.font.underline,
                                     strike=source_cell.font.strike,
                                     color=source_cell.font.color)

            target_cell.border = Border(left=source_cell.border.left,
                                         right=source_cell.border.right,
                                         top=source_cell.border.top,
                                         bottom=source_cell.border.bottom)

            # Используем PatternFill вместо Fill
            if source_cell.fill:
                target_cell.fill = PatternFill(start_color=source_cell.fill.start_color,
                                                 end_color=source_cell.fill.end_color,
                                                 fill_type=source_cell.fill.fill_type)

            target_cell.number_format = source_cell.number_format
            target_cell.protection = Protection(locked=source_cell.protection.locked,
                                                 hidden=source_cell.protection.hidden)
            target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                               vertical=source_cell.alignment.vertical,
                                               text_rotation=source_cell.alignment.text_rotation,
                                               wrap_text=source_cell.alignment.wrap_text,
                                               shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                               indent=source_cell.alignment.indent)

# ok
@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_current_user_info(request):
    user = request.user
    queryset = CustomUser.objects.get(id=user.id)
    serializer_data = CustomUserSerializer(queryset, many=False).data
    return Response(serializer_data)

# PROVERKA
class CourtsListAPI(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        active = request.GET.get('active', False)

        queryset = CourtCases.objects.filter(archive=active)
        if not (user.chief_rule or user.is_admin):
            queryset = queryset.filter(user=user.id)

        queryset = queryset.annotate(
            date_nearest=Max('events__dates_of_court_hearing'),
        )

        # Фильтрация
        filters = request.GET.get('filters', None)
        if filters:
            filters = json.loads(filters).get('items', [])
            queryset = filter_queryset(queryset, filters)

        # Сортировка
        sort_by = request.GET.get('sortBy', None)
        sort_type = request.GET.get('sortType', 'asc')
        if sort_by:
            queryset = queryset.order_by(f'-{sort_by}' if sort_type == 'desc' else sort_by)

        # Пагинация
        paginator = CustomPagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        # Сериализация данных
        serializer_data = CourtCasesSerializer(paginated_queryset, many=True).data

        # Возвращаем ответ
        return paginator.get_paginated_response(serializer_data)

    def put(self, request):
        data = request.data
        try:
            courts = CourtCases.objects.filter(id__in=data['list_ids'])
            courts.update(archive=data['flag'])
            events = EventsCourtCases.objects.filter(court_case__in=data['list_ids']).values('court_case').annotate(max_id=Max('id'))
            for event in events:
                print(f"Для дела {event['court_case']}: ", create_task(event['max_id']))
            return Response("Дела успешно изменены", status=status.HTTP_200_OK)
        except Exception as e:
            print("ERROR - CourtsListAPI - put: ", e)
            return Response(f'Не удалось загрузить дела: {str(e)}', status=status.HTTP_400_BAD_REQUEST)
# PROVERKA
class CourtSelfAPI(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        queryset = CourtCases.objects.get(id=pk)

        serializer_class = CourtCasesSerializer(queryset, many=False)

        return Response(serializer_class.data)

    def post(self, request):
        user = request.user
        
        # Используем сериализатор для валидации
        serializer = CourtCasesSerializer(data={**request.data, "user": user.id})
        
        if not serializer.is_valid():
            # Возвращаем детальные ошибки валидации
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        court = serializer.save(is_checked=True)
        
        return Response({'id': court.id}, status=status.HTTP_201_CREATED)
            

    def put(self, request, pk):
        court = CourtCases.objects.get(id=pk)
        data = request.data.copy()  # Создаем копию чтобы не менять оригинал
        
        # Извлекаем events до валидации основного сериализатора
        events_data = data.pop('events', [])
        
        # Обновляем основную модель
        serializer = CourtCasesSerializer(court, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        max_type_event = 0
        max_event_id = None

        if events_data:
            event_ids = [event_data.get('id') for event_data in events_data if event_data.get('id')]
            events = EventsCourtCases.objects.filter(id__in=event_ids).select_related('type_event')
            events_dict = {event.id: event for event in events}
            
            for event_data in events_data:
                event_id = event_data.get('id')
                if event_id and event_id in events_dict:
                    event = events_dict[event_id]
                    event_serializer = EventsCourtCasesSerializer(event, data=event_data, partial=True)
                    event_serializer.is_valid(raise_exception=True)
                    event_serializer.save()
                    
                    if event.type_event.id > max_type_event:
                        max_type_event = event.type_event.id
                        max_event_id = event_id

        if max_event_id:
            print(create_task(max_event_id))

        return Response(serializer.data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        user = request.user

        court = CourtCases.objects.get(id=pk)

        serializer_class = CourtCasesSerializer(court, many=False)
        
        if user.is_admin is False and serializer_class.data['user'] != user.id:
            return Response(f'У вас нет прав доступа на удаление данного дела.', status=status.HTTP_400_BAD_REQUEST)
        else:
            court.delete()
            return Response("Дело удалено", status=status.HTTP_200_OK)

class EventSelfAPI(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            data = request.data
            type_event = TypesEventsCourtCases.objects.get(id=data['type_event'])
            court = CourtCases.objects.get(id=pk)
            event = EventsCourtCases.objects.create(type_event=type_event, court_case=court)
            create_task(event.id)
            serializer_class = EventsCourtCasesSerializer(event, many=False)
            return Response(serializer_class.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            print('EventSelfAPI - post - error:', e)
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)
      
    def delete(self, request, pk):
        try:
            court = CourtCases.objects.get(id=pk)
            event = EventsCourtCases.objects.filter(court_case=court).order_by('-id').first()
            event.delete()
            return Response(status=status.HTTP_200_OK)
        except Exception as e:
            print('EventSelfAPI - delete - error:', e)
            return Response(status=status.HTTP_400_BAD_REQUEST)

class NotifiesAPI(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        
        queryset = NotifyTask.objects.filter(target=user.id)

        # Фильтрация
        filters = request.GET.get('filters', None)
        if filters:
            filters = json.loads(filters).get('items', [])
            queryset = filter_queryset(queryset, filters)

        # Сортировка
        sort_by = request.GET.get('sortBy', None)
        sort_type = request.GET.get('sortType', 'asc')
        if sort_by:
            queryset = queryset.order_by(f'-{sort_by}' if sort_type == 'desc' else sort_by)

        # queryset

        # Пагинация
        paginator = CustomPagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        # Сериализация данных
        serializer_data = NotifyTasksSerializer(paginated_queryset, many=True).data

        return paginator.get_paginated_response(serializer_data)
    
    def put(self, request):
        keys = request.data['keys']
        queryset = NotifyTask.objects.filter(pk__in=keys)

        queryset.delete()

        return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def create_excel(request):
    user = request.user
    
    queryset = CourtCases.objects.filter(archive=False)
    if not (user.chief_rule or user.is_admin):
        queryset = queryset.filter(user=user.id)

    queryset = queryset.annotate(
        date_nearest=Max('events__dates_of_court_hearing'),
    )

    # Фильтрация
    filters = request.GET.get('filters', None)
    if filters:
        filters = json.loads(filters).get('items', [])
        queryset = filter_queryset(queryset, filters)
    
    template = DocTemplate.objects.get(system='UO', name='court_out')
    with open(template.file.path, 'rb') as file:
        excel_data = io.BytesIO(file.read())

    data = CourtCasesSerializer(queryset, many=True).data
    wb = load_workbook(filename=excel_data, read_only=False)
    ws = wb.active
    ws['A2'] = user.name

    # заполнение фильтров 
    if filters:
        for indx, item in enumerate(filters):
            field_name = item.get('field')
            value = item.get('value')
            field = CourtCases._meta.get_field(field_name)
            verbose_name = field.verbose_name
            col = indx+1
            ws.cell(row=5, column=col).value = verbose_name
            ws.cell(row=6, column=col).value = value
    else:
        ws.cell(row=5, column=1).value = "Фильтры не применялись"

    row_indx = 12
    for item in data:
        ws.cell(row=row_indx, column=1).value = item['number_case_in_numenklature']
        ws.cell(row=row_indx, column=2).value = item['number_case_in_first_instance']
        ws.cell(row=row_indx, column=3).value = item['name_court']
        ws.cell(row=row_indx, column=4).value = item['plaintiff']
        ws.cell(row=row_indx, column=5).value = item['defendant']
        ws.cell(row=row_indx, column=6).value = item['category']
        ws.cell(row=row_indx, column=7).value = item['claim']
        ws.cell(row=row_indx, column=8).value = item['claim_summ']
        ws.cell(row=row_indx, column=9).value = '\n'.join([notify['message'] for notify in item['notifyes']])
        
        if item.get('events'):
            for indx, event in enumerate(item['events']):
                col = 10 + (indx * 5)
                ws.cell(row=row_indx, column=col).value = event['date_of_dicision'] if event['type_event'] == 1 else event['date_appel_issue']
                ws.cell(row=row_indx, column=col + 1).value = event['date_of_dicision_force'] if event['type_event'] == 1 else event['dates_of_court_hearing']
                ws.cell(row=row_indx, column=col + 2).value = event['brief_operative_part']
                ws.cell(row=row_indx, column=col + 3).value = 'Да' if event['need_appel'] else 'Нет'
                if event["date_minfin"] or event["number_letter_minfin"]:
                    ws.cell(row=row_indx, column=col + 4).value = f'{event["date_minfin"]} | {event["number_letter_minfin"]}' 
        copy_row(ws, row_indx)
    
    ws.delete_rows(row_indx)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'inline; filename="Выгрузка юридических дел.xlsx"'

    return response

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def load_excel(request):
    try:
        user = request.user
        excel_file = request.FILES['file']
        wb = load_workbook(excel_file)
        ws = wb.active
        count_created = 0
        total_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_count += 1 
            find_court = CourtCases.objects.filter(number_case_in_numenklature=row[0])
            if not find_court:
                count_created += 1
                sum_claim = 0.0 if not row[6] else float(row[6].replace('\xa0', '').replace(',', '.').strip()) 
                CourtCases.objects.create(
                    number_case_in_numenklature=row[0],
                    number_case_in_first_instance=row[2] or "Выгрузка из СКИАО", 
                    plaintiff=row[3], 
                    defendant=row[4], 
                    category=row[5],
                    claim_summ=sum_claim, 
                    user=user.id,
                    name_court="Выгрузка из СКИАО",
                    is_loaded=True
                )

        return Response({"total_count": total_count, "count_created": count_created})
    except Exception as e:
        print(e)
        return Response('Ошибка чтения файла', status=status.HTTP_400_BAD_REQUEST)

