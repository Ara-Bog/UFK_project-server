from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from .models import *
from django.db.models import Q, Max
from .serializers import *
from django.contrib.postgres.search import SearchRank, SearchQuery, SearchVector
from openpyxl import load_workbook
import io
import datetime

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def search(request):
    query = request.data.get('query')
    if not query:
        return Response("Нет запроса", status=400)

    user_vector = SearchVector('name', 'job__name')
    users = CustomUser.objects.select_related('job').annotate(
        rank=SearchRank(user_vector, SearchQuery(query))
    ).filter(
        Q(name__icontains=query) | Q(email__icontains=query) | Q(job__name__icontains=query)
    )

    dep_vector = SearchVector('name')
    deps = Departments.objects.annotate(
        rank=SearchRank(dep_vector, SearchQuery(query))
    ).filter(
        Q(name__icontains=query)
    )

    results = list(users) + list(deps)
    results.sort(key=lambda x: getattr(x, 'rank', 0), reverse=True)
    results = [
        {
            'id': x.id,
            'name': x.name,
            'email': getattr(x, 'email', None),
            'job__str': x.job.name if hasattr(x, 'job') else None,
            'type': 'user' if isinstance(x, CustomUser) else 'department',
            **({
                'isFired': x.isFired,
                'isDisabled': x.isDisabled
            } if isinstance(x, CustomUser) else {})
        }
        for x in results
    ]

    return Response(results, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def departmentInfo(request, pk):
    users_data = CustomUser.objects.filter(department__id=pk)
    department_name = Departments.objects.get(id=pk).name
    return Response(data={'users': UserSerializer(users_data, many=True).data, 'name': department_name}, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def userInfo(request, pk):
    try:
        user_data = UserSerializer(CustomUser.objects.get(id=pk)).data
        user_roles_data = UserRoles.objects.filter(
            user=pk, isChecked=True
        ).select_related('role', 'system', 'system__parent').values(
            'system_id', 'system__name', 'system__parent_id',
            'role__name', 'role_id'
        )
        
        if not user_roles_data:
            return Response({
                'roles': [],
                'user': user_data
            })
        
        updates_data = LogUpdates.objects.filter(user=pk).values(
            'system_id'
        ).annotate(
            last_update=Max('date_msg')
        ).values('system_id', 'last_update')
        
        updates_dict = {item['system_id']: item['last_update'] for item in updates_data}
        
        systems_dict = {}
        for item in user_roles_data:
            system_id = item['system_id']
            system_name = item['system__name']
            parent_id = item['system__parent_id']
            role_name = item['role__name']
            
            if parent_id is None:
                if system_id not in systems_dict:
                    systems_dict[system_id] = {
                        'id': system_id,
                        'name': system_name,
                        'roles': [],
                        'subsystems': [],
                        'last_update': updates_dict.get(system_id)
                    }
                systems_dict[system_id]['roles'].append({'name': role_name})
            else:
                if system_id not in systems_dict:
                    systems_dict[system_id] = {
                        'id': system_id,
                        'name': system_name,
                        'parent_id': parent_id,
                        'roles': [],
                        'last_update': updates_dict.get(system_id),
                        'is_subsystem': True
                    }
                systems_dict[system_id]['roles'].append({'name': role_name})
                
                if parent_id not in systems_dict:
                    parent_system = Systems.objects.filter(id=parent_id).values('id', 'name').first()
                    if parent_system:
                        systems_dict[parent_id] = {
                            'id': parent_id,
                            'name': parent_system['name'],
                            'roles': [],
                            'subsystems': [],
                            'last_update': updates_dict.get(parent_id)
                        }
        
        result = []
        for system in systems_dict.values():
            if system.get('is_subsystem'):
                parent_id = system['parent_id']
                if parent_id in systems_dict:
                    subsystem_data = system.copy()
                    del subsystem_data['is_subsystem']
                    del subsystem_data['parent_id']
                    systems_dict[parent_id]['subsystems'].append(subsystem_data)
            else:
                result.append(system)
        
        return Response({
            'roles': result,
            'user': user_data
        })
        
    except Exception as e:
        return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def rolesUserToExcel(request, pk):
    try:
        template = DocTemplate.objects.get(system='MATRIX', name='roles_user')
        with open(template.file.path, 'rb') as file:
            excel_data = io.BytesIO(file.read())

        wb = load_workbook(excel_data, read_only=False)
        ws = wb["Данные"]

        user_data = CustomUser.objects.select_related('department', 'job').get(id=pk)
        roles_queryset = UserRoles.objects.filter(
            user=pk, 
            isChecked=True 
        ).select_related(
            'system', 'system__parent', 'role' 
        ).order_by('system__parent__name', 'system__name')

        ws['B3'] = user_data.department.name if user_data.department else ''
        ws['B4'] = user_data.name
        ws['B5'] = user_data.job.name if user_data.job else ''
        ws['B6'] = datetime.datetime.now().strftime("%d.%m.%Y")

        cur_system = None
        cur_subsystem = None
        
        for i, user_role in enumerate(roles_queryset):
            cur_row = 9 + i

            select_system = user_role.system.parent.name if user_role.system.parent else user_role.system.name
            select_subsystem = user_role.system.name if user_role.system.parent else None

            if cur_system != select_system:
                ws[f'A{cur_row}'] = select_system
                ws[f'B{cur_row}'] = select_subsystem
                cur_system = select_system
                cur_subsystem = select_subsystem
            elif cur_subsystem != select_subsystem:
                ws[f'B{cur_row}'] = select_subsystem
                cur_subsystem = select_subsystem

            ws[f'E{cur_row}'] = user_role.role.name
            
            if i:
                ws.merge_cells(f'B{cur_row}:D{cur_row}')
                ws.merge_cells(f'E{cur_row}:H{cur_row}')
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                cell = ws[f'{col}{cur_row}']
                if ws['A9']._style:
                    cell._style = ws['A9']._style

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename="Данные по сотруднику.xlsx"'
        return response

    except DocTemplate.DoesNotExist:
        return Response("Шаблон не найден", status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(f"Error processing template file: {e}")
        return Response("Ошибка в формировании файла", status=status.HTTP_409_CONFLICT)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def rolesDepartmentToExcel(request, pk):
    try:
        template = DocTemplate.objects.get(system='MATRIX', name='roles_department')
        with open(template.file.path, 'rb') as file:
            excel_data = io.BytesIO(file.read())

        wb = load_workbook(excel_data, read_only=False)
        ws = wb["Данные"]
        table = ws.tables['Main']

        users_queryset = CustomUser.objects.filter(department=pk)
        roles_queryset = UserRoles.objects.filter(
            user__in=users_queryset.values_list('id', flat=True),
            isChecked=True  
        ).select_related('user', 'system', 'system__parent', 'role')

        department_name = Departments.objects.get(id=pk).name
        ws['A1'] = department_name

        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)

        for user_role in roles_queryset:
            parent = user_role.system.parent  
            main_system = parent.name if parent else user_role.system.name
            subsystem = user_role.system.name if parent else None

            ws.append([
                user_role.user.name, 
                main_system,
                subsystem,  
                user_role.role.name,  
            ])

        if ws.max_row >= 3:
            table.ref = f"A2:D{ws.max_row}"
        else:
            ws.append(["", "", "", ""])
            table.ref = "A2:D3"
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = f'attachment; filename="Данные по отделу {department_name}.xlsx"'
        return response

    except DocTemplate.DoesNotExist:
        return Response("Шаблон не найден", status=status.HTTP_404_NOT_FOUND)
    except Departments.DoesNotExist:
        return Response("Отдел не найден", status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(f"Error processing template file: {e}")
        return Response("Ошибка в формировании файла", status=status.HTTP_409_CONFLICT)