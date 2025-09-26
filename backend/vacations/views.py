from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from .serializers import *
from .models import *
from rest_framework.views import APIView
from datetime import datetime, timedelta
from openpyxl import load_workbook
from django.http import HttpResponse
import os
import io
import zipfile
from holidays import country_holidays
import json
from django.db.models import Q, Prefetch, Min
from django.db import transaction
from Auth_LDAP.models import CustomUser
from rest_framework.authentication import TokenAuthentication
from mui_table_settings import filter_queryset, CustomPagination, get_ordering_field
from consts import MIME_TYPES
from urllib.parse import quote
from Auth_LDAP.views import generate_document_content

def set_default_settings(user: CustomUser):
    new_settings = UserSettings.objects.create(
        user = user,
        manager = CustomUser.objects.filter(job__name="Руководитель").first(),
        chief = CustomUser.objects.filter(job__name="Начальник отдела", department=user.department).first(),
        period = datetime.now().year
    )

    return new_settings

HOLIDAYS = country_holidays('RU', observed=False, language="ru")

def check_date(start:str, end:str):
    dates = HOLIDAYS[start:end]
    counter = 0
    for day in dates:
        if "Выходной" not in HOLIDAYS.get(day):
            counter += 1
    return counter

def calc_date_end(date_start:str, court_days:int):
    date_start_obj = datetime.strptime(date_start, '%d.%m.%Y')
    date_start_str = date_start_obj.strftime('%Y-%m-%d')
    date_end = date_start_obj + timedelta(days=court_days-1)
    skip_days = -1
    accum = 1
    while skip_days:
        accum += skip_days
        skip_days = check_date(date_start_str, (date_end + timedelta(days=1)).strftime('%Y-%m-%d'))
        skip_days -= accum
        date_end = date_end + timedelta(days=skip_days)

    return date_end.strftime('%d.%m.%Y')

def check_vacation_conflicts(target_user: CustomUser, new_start_str: str, new_end_str: str, exclude_ids=[]):
    """
    Проверяет пересечение нового отпуска с существующими отпусками сотрудника
    Возвращает Queryset
    """

    new_start = datetime.strptime(new_start_str, '%d.%m.%Y').date()
    new_end = datetime.strptime(new_end_str, '%d.%m.%Y').date()

    conflicts = Vacations.objects.filter(
        user=target_user,
        is_archive=False,
        is_transfered=False,
    ).exclude(id__in=exclude_ids).filter(
        Q(date_start__range=(new_start, new_end)) |
        Q(date_end__range=(new_start, new_end)) |
        Q(date_start__lte=new_start, date_end__gte=new_end) |
        Q(date_start__gte=new_start, date_end__lte=new_end)
    )
    
    return conflicts


@api_view(['POST'])
def test(request):
    return Response(BaseVacationSerializer(Vacations.objects.all(), many=True).data, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_current_user_info(request):
    return Response(CustomUserSerializer(request.user).data)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_responsibles(request):
    queryset = CustomUser.objects.filter(OGGSK=True, isChecked=True)
    serializer_data = CustomUserSerializer(queryset, many=True).data
    return Response(serializer_data)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_managers(request):
    queryset = CustomUser.objects.filter(isChecked=True, manager_rule=True)
    serializer_data = CustomUserSerializer(queryset, many=True).data
    return Response(serializer_data)

class UserSettingsAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        try:
            settings = UserSettings.objects.get(user=user.id)
        except UserSettings.DoesNotExist:
            settings = set_default_settings(user)

        serializer = UserSettingsWriteSerializer(settings)
        return Response(serializer.data)
    
    def post(self, request):
        user = request.user
        data = request.data
        
        try:
            settings = UserSettings.objects.get(user=user)
        except UserSettings.DoesNotExist:
            settings = set_default_settings(user)

        serializer = UserSettingsWriteSerializer(settings, data=data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()

        return Response(serializer.data, status=status.HTTP_200_OK)

class ListVacationsAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period = UserSettings.objects.get(user=request.user).period

        # Начинаем формировать запрос
        queryset = Vacations.objects.filter(is_archive=False, is_transfered=False, date_start__year=period).order_by('-id')

        # Фильтрация
        filters = request.GET.get('filters', None)
        if filters:
            filters = json.loads(filters).get('items', [])
            queryset = filter_queryset(queryset, filters)

        # Сортировка
        sort_by = request.GET.get('sortBy', None)
        sort_type = request.GET.get('sortType', 'asc')
        if sort_by:
            ordering_field = get_ordering_field(queryset.model, sort_by)
            if sort_type == 'desc':
                ordering_field = f'-{ordering_field}'
            queryset = queryset.order_by(ordering_field)

        # Пагинация
        paginator = CustomPagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        serializer_data = VacationListSerializer(paginated_queryset, many=True).data

        return paginator.get_paginated_response(serializer_data)

class FillVacationsAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def convert_date_format(self, date_string, today):
        date_format = "%d.%m.%Y"
        new_date_format = "%Y-%m-%d"
        date_object = datetime.strptime(date_string, date_format)
        if date_object.date() < today:
            return None
        new_date_string = date_object.strftime(new_date_format)
        return new_date_string
    
    def get(self, request):
        try:
            template = DocTemplate.objects.get(system='OGGSK', name='form_vacations')
            original_filename = template.file.name
            _, file_extension = os.path.splitext(original_filename)
            download_filename = f"{template.name_visible}{file_extension}"
            quoted_filename = quote(download_filename)

            with open(template.file.path, 'rb') as file:
                excel_data = io.BytesIO(file.read())
                
            wb = load_workbook(
            filename=excel_data, read_only=False, keep_vba=True)

            ws_jobs = wb["Должности"]
            ws_departaments = wb["Отделы"]
            ws_workers = wb["Сотрудники"]

            table_jobs = ws_jobs.tables['Jobs']
            table_departaments = ws_departaments.tables['Departaments']
            table_workers = ws_workers.tables['Workers']

            queryset_jobs = Jobs.objects.all().exclude(sortBy=100).order_by('sortBy')
            queryset_departments = Departments.objects.all().exclude(sortBy=100).order_by('sortBy')
            queryset_workers = CustomUser.objects.filter(
                is_superuser=False,
                tab_number__isnull=False,
                isDisabled=False,
                isFired=False,
            ).filter(
                Q(isChecked=True) | Q(manually_added=True)
            ).order_by('name')
            
            for item in queryset_jobs:
                ws_jobs.append([item.name, item.sortBy])
            ws_jobs.delete_rows(2, 1)
            table_jobs.ref = "A1:B" + str(ws_jobs.max_row)

            for item in queryset_departments:
                ws_departaments.append([item.name, item.sortBy])
            ws_departaments.delete_rows(2, 1)
            table_departaments.ref = "A1:B" + str(ws_departaments.max_row)

            for item in queryset_workers:
                ws_workers.append([item.name, item.tab_number, item.department.name, item.job.name])
            ws_workers.delete_rows(2, 1)
            table_workers.ref = "A1:D" + str(ws_workers.max_row)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            http_response = HttpResponse(
                output,
                status=status.HTTP_200_OK,
                content_type=MIME_TYPES.get(file_extension.lower(), 'application/octet-stream')
            )
            http_response['Content-Disposition'] = f'attachment; filename="{quoted_filename}"'
            return http_response
            
        except DocTemplate.DoesNotExist:
            return Response(
                "Шаблон не найден", 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                f"Ошибка обработки файла: {str(e)}", 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request):
        user = request.user
        excel_file = request.FILES['file']
        total = 0
        suc = 0
        cross = 0
        repeat = 0
        ovner = 0
        err = 0
        data = {
            "user": None,
            "count_days": None,
            "date_start": None,
            "date_end": None,
            "user_created": str(user)
        }
        try:
            today = datetime.today().date()
            wb = load_workbook(excel_file)
            range_table = wb["Base"].tables["Main"].ref
            tab_user = None
            current_user = None
            for row in wb["Base"][range_table][1:]:
                total += 1
                current_tab = row[0].value
                if tab_user != current_tab:
                    current_user = CustomUser.objects.get(tab_number=current_tab)
                    tab_user = current_tab
                    data.update({
                        "user": current_user.id
                    })
                try:
                    new_date = datetime.strptime(row[2].value, '%d.%m.%Y').date()
                    format_date = new_date.strftime("%d.%m.%Y")
                    date_end = None
                    if today > new_date:
                        ovner += 1
                        continue
                    else:
                        date_end = calc_date_end(format_date, row[1].value)
                    check_repeat = Vacations.objects.filter(user=current_user, date_start=new_date, is_archive=False, is_transfered=False).count()
                    if check_repeat:
                        repeat += 1
                        continue
                    check_conflict = check_vacation_conflicts(current_user, format_date, date_end)
                    if check_conflict.exists():
                        cross += 1
                        continue
                    data.update({
                        "count_days": row[1].value,
                        "date_start": format_date,
                        "date_end": date_end,
                    })
                    serializer = BaseVacationSerializer(
                        data=data,
                        context={'request': request, 'create_by': 'load'}
                    )
                    if not serializer.is_valid():
                        raise serializers.ValidationError(serializer.errors)
                    serializer.save()
                    suc += 1
                except Exception as e:
                    print("Ошибка обработка строки: ", e)
                    err += 1
            wb.close()
        except (zipfile.BadZipFile, OSError) as e:
            print(f"Ошибка чтения файла: {e}")
            return Response("Не верный формат файла", status=status.HTTP_409_CONFLICT)
        except Exception as e:
            print(f"Ошибка обработки строк: {e}")
            return Response("Непредвиденная ошибка", status=status.HTTP_409_CONFLICT)
        
        return Response(f"Найдено отпусков - {total}; Успешно загружено - {suc};Пересекающиеся - {cross}; Повторы - {repeat}; Пропущено - {ovner}; Ошибка обработки строк - {err}", 
                        status=status.HTTP_201_CREATED)

class VacationAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        vac = Vacations.objects.prefetch_related(
            Prefetch('drop_vacations', queryset=TransactionTransferVacations.objects.only('id')),
            Prefetch('add_vacations', queryset=TransactionTransferVacations.objects.only('id'))
        ).get(id=pk)

        serializer_vac = VacationSerializer(vac).data
        
        logs_qery = Logs.objects.filter(vacation=vac).order_by('-data_create', '-time_create')
        serializer_logs = LogsSerializer(logs_qery, many=True).data
        return Response({'vacation': serializer_vac, 'logs': serializer_logs})
    
    def post(self, request, pk):
        user = request.user

        vacation = Vacations.objects.get(id=pk)
        vacation.responsible = user
        vacation.save()

        return Response(status=status.HTTP_200_OK)

    def put(self, request, pk):
        data = request.data
        vacation = Vacations.objects.get(id=pk)
        data['responsible_id'] = data.get('responsible', {}).get('id', None)
        ser = VacationSerializer(vacation, data=data, partial=True,
        context={'request': request})
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)

        ser.save()
        queryset = Logs.objects.filter(vacation_id=vacation).order_by('-data_create', '-time_create')
        serializer_logs = LogsSerializer(queryset, many=True).data

        return Response(serializer_logs, status=status.HTTP_200_OK)

class TransferAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        transaction = TransactionTransferVacations.objects.get(id=pk)
        transaction_data = TransfersSerializer(transaction).data

        return Response(transaction_data, status=status.HTTP_200_OK)
    
    @transaction.atomic
    def post(self, request):
        user = request.user
        data = request.data

        target = data.get('select_user')
        vacations_drop_ids = data.get('vacations_drop', [])
        vacations_create_data = data.get('vacations_create', [])

        # Получаем все отпуска сотрудника (кроме переносимых)
        existing_vacations = Vacations.objects.filter(
            user=target,
            is_transfered=False,
            is_archive=False,
        ).exclude(id__in=vacations_drop_ids).order_by('date_start')

        for vacation_data in vacations_create_data:
            new_start = datetime.strptime(vacation_data['date_start'], '%d.%m.%Y').date()
            new_end = datetime.strptime(calc_date_end(vacation_data['date_start'], vacation_data['count_days']), '%d.%m.%Y').date()

            vacation_data['date_start'] = new_start
            vacation_data['date_end'] = new_end

        temp_vacations = sorted(vacations_create_data, key=lambda x: x['date_start'])
        check_vacations = [temp_vacations[0]]

        vacations_create = []
        vacations_merged = [] 
        for vacation_data in temp_vacations:
            # проверяем корректность переданных данных. отпуска не должны пересекатся
            if len(check_vacations) > 1:
                last_vacation = check_vacations[-1]
                if vacation_data['date_start'] <= last_vacation['date_end']:
                    transaction.set_rollback(True)
                    return Response("Новые отпуска пересекаются между собой. Проверьте коррекность ввода.", status=status.HTTP_400_BAD_REQUEST)

            # Объединение существующих отпусков с новыми в один (при разнице в 1 день)
            for existing in existing_vacations:
                existing_start = existing.date_start
                existing_end = existing.date_end
                
                if (existing_end + timedelta(days=1) == new_start) or (new_end + timedelta(days=1) == existing_start):
                    vacation_data['date_start'] = min(existing_start, vacation_data['date_start'])
                    vacation_data['date_end'] = max(existing_end, vacation_data['date_end'])
                    vacation_data['count_days'] += existing.count_days
                    
                    serializer = BaseVacationSerializer(
                        instance=existing,
                        data={
                            'is_transfered': True,
                            'is_merged': True
                        },
                        partial=True,
                        context={'request': request}
                    )
                    if serializer.is_valid():
                        existing_vacations.exclude(id=existing.id)
                        serializer.save()
                        vacations_merged.append(existing.id)
                    else:
                        transaction.set_rollback(True)
                        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            check_vacations.append(vacation_data)
            
        # сортировка для оптимизации
        temp_vacations = sorted(vacations_create_data, key=lambda x: x['date_start'])
        check_vacations = [temp_vacations[0]]
        # проверяем пересечения новых отпусков между собой для объединения
        for vacation in temp_vacations[1:]:
            last_vacation = check_vacations[-1]
            if last_vacation['date_start'] <= vacation['date_start'] <= last_vacation['date_end']:
                last_vacation['count_days'] = last_vacation['count_days'] + (vacation['date_end'] - last_vacation['date_end']).days
                last_vacation['date_end'] = vacation['date_end']
            else:
                last_vacation['date_start'] = last_vacation['date_start'].strftime('%d.%m.%Y')
                last_vacation['date_end'] = last_vacation['date_end'].strftime('%d.%m.%Y')
                check_vacations.append(vacation)
        
        check_vacations[-1]['date_start'] = check_vacations[-1]['date_start'].strftime('%d.%m.%Y')
        check_vacations[-1]['date_end'] = check_vacations[-1]['date_end'].strftime('%d.%m.%Y')

        # проверяем наличие конфликтов с существующими и создаем новые отпуска
        for vacation_data in check_vacations:
            conflicts = check_vacation_conflicts(target, vacation_data['date_start'], vacation_data['date_end'], vacations_drop_ids)
            if conflicts.exists():
                transaction.set_rollback(True)
                return Response(
                    "Новые отпуска пересекаються с уже имеющимеся. Проверьте корректность данных.",
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            vacation_data.update({
                'user': target,
                'user_created': str(request.user),
                'create_by_transfered': True,
            })
            serializer = BaseVacationSerializer(data=vacation_data, context={'request': request, 'create_by': 'create_by_transfered'})
            if serializer.is_valid():
                vacation = serializer.save()
                vacations_create.append(vacation)
            else:
                transaction.set_rollback(True)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Создаем транзакцию переноса
        transaction_obj = TransactionTransferVacations.objects.create(
            user_create=user
        )

        vacations_drop = Vacations.objects.filter(id__in=vacations_drop_ids)
        
        vacations_merged = Vacations.objects.filter(id__in=vacations_merged)

        # Добавляем отпуска в транзакцию
        transaction_obj.vacations_drop.set(vacations_drop | vacations_merged)
        transaction_obj.vacations_create.set(vacations_create)
        
        # Обновляем флаг is_transfered для исходных отпусков
        vacations_drop.update(is_transfered=True)
        
        # Сериализуем результат
        serializer = TransfersSerializer(transaction_obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_userifno_doc(request, pk):
    target = CustomUser.objects.get(id=pk)
    serialize_data_user = CustomUserSerializer(target).data
    period = UserSettings.objects.get(user=request.user).period
    
    vacations_query = Vacations.objects.filter(user=target, is_archive=False, is_transfered=False, date_start__year=period).order_by('date_start')
    serialize_data_vacations = VacationSerializer(vacations_query, many=True).data

    return Response({'user':serialize_data_user, 'vacations': serialize_data_vacations})


class UserPageAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        user = CustomUser.objects.get(id=pk)
        serialize_data_user = CustomUserSerializer(user).data
        
        vacations_query = Vacations.objects.filter(user=user).order_by('is_transfered', 'is_archive', 'date_start')
        serialize_data_vacations = VacationSerializer(vacations_query, many=True).data

        return Response({'user':serialize_data_user, 'vacations': serialize_data_vacations})
    
    def put(self, request, pk):
        data = request.data

        user = CustomUser.objects.get(id=pk)
        
        ser = CustomUserSerializer(user, data=data, partial=True)
        if ser.is_valid():
            ser.save()
        else:
            return Response(f'Ошибка сохранения данных пользователя: {ser.errors}', status=status.HTTP_400_BAD_REQUEST)
        
        return Response(ser.data)

class UsersAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Получаем параметры запроса
        sort_by = request.GET.get('sortBy', None) 
        sort_type = request.GET.get('sortType', 'asc')  
        filters = json.loads(request.GET.get('filters', '{}')).get('items', [])  
        
        users_query = CustomUser.objects.filter(
            Q(isChecked=True) | Q(manually_added=True)
        ).annotate(
            date_nearest=Min(
                'uservacations_as_user__date_start',
                filter=Q(uservacations_as_user__is_archive=False, uservacations_as_user__is_transfered=False)
            )
        ).order_by('-id')

        # Применяем фильтры
        users_query = filter_queryset(users_query, filters)


        if sort_by:
            ordering_field = get_ordering_field(users_query.model, sort_by)
            if sort_type == 'desc':
                ordering_field = f'-{ordering_field}'
            users_query = users_query.order_by(ordering_field)

        # Пагинация
        paginator = CustomPagination()
        paginated_users = paginator.paginate_queryset(users_query, request)

        # Сериализация данных
        serializer_data = CustomUserSerializer(paginated_users, many=True).data

        # Возвращаем ответ
        return paginator.get_paginated_response(serializer_data)

class UserPeriodsAPIView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk):
        periods_list = PeriodsVacation.objects.filter(user=pk).order_by('period_start', 'period_end')
        serializer_data = PeriodsSerializer(periods_list, many=True).data
        return Response(serializer_data)

    def post(self, request, pk):
        user = CustomUser.objects.get(id=pk)
        serializer = PeriodsSerializer(data=request.data, many=True, context={'user': user})
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Удаляем старые и создаем новые периоды
        PeriodsVacation.objects.filter(user=user).delete()
        serializer.save()
        
        return self.get(request, pk)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def check_correct_tab_number(request):
    count = CustomUser.objects.filter(Q(is_superuser=False, tab_number__isnull=True), Q(isChecked=True) | Q(manually_added=True)).count()
    return Response({'flag': count == 0}, status=status.HTTP_202_ACCEPTED)

@api_view(['PUT'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def take_to_work(request):
    user = request.user
    data = request.data['list_ids']
    Vacations.objects.filter(id__in=data).update(responsible=user)
    return Response(status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def upload_tab_number(request):
    excel_file = request.FILES['file']
    total = 0
    suc = 0
    err = 0
    skip = 0
    try:
        wb = load_workbook(excel_file)
        worksheet = wb[wb.sheetnames[0]]
        for row in list(worksheet.rows)[1:]:
            total += 1
            try:
                user = [i.value.lower() for i in row[:3]]
                name_str = ''.join(user)
                current_user = CustomUser.objects.get(name_key=name_str)
                if current_user.tab_number == row[3].value:
                    skip += 1
                    continue
                current_user.tab_number = row[3].value
                current_user.save()
                suc += 1
            except Exception as e:
                err += 1
                print(f'Error load tab_number user: {" ".join(user)}')
        
        wb.close()
        response_message = f'Всего обработано: {total}; Обработано успешно: {suc}; Ошибка в обработке: {err}; Не требуется обновление: {skip}'
        return Response(response_message, status=status.HTTP_201_CREATED)
    except Exception as e:
        print(f"Error processing file: {e}")
        return Response("Не верный формат файла", status=status.HTTP_409_CONFLICT)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def load_plan(request):
    vacations_data = Vacations.objects.filter(is_archive=False, is_transfered=False).select_related(
        'user__department', 
        'user__job'
    ).order_by(
        'user__department__sortBy',
        'user__job__sortBy', 
        'user__name',
        'date_start'
    )

    serializer_data = VacationListSerializer(vacations_data, many=True).data
    settings = UserSettingsReadSerializer(UserSettings.objects.get(user=request.user.id)).data
    try:
        result = generate_document_content(
            dataset={
                'vacations': serializer_data,
                'settings': settings,
                'doc_settings': {'variant': 'plan'}
            },
            type_doc='plan_grafic',
            system='OGGSK',
            doc_for=''
        )
        
        response = HttpResponse(
            result['content'],
            content_type=result['content_type']
        )
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quote(result["filename"])}'
        return response
        
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def load_docs_employee(request):
    
    type_doc = request.data.get('type_doc', None)
    vacation_id = request.data.get('vacation_id', None)
    added_users_ids = request.data.get('added_users_ids', None)
    target = request.data.get('target', None)

    vacation = None
    settings = {}
    data = {
        'doc_settings': request.data.get('doc_settings', None),
        'params': request.data.get('params', {})
    }

    headers = {
        'Authorization': f"Token {request.headers.get('Authorization')}",
    }

    if vacation_id:
        vacation = Vacations.objects.get(id=vacation_id)
        data['vacation'] = VacationSerializer(vacation).data
        if data['vacation']['periods']:
            periods_vac = PeriodsVacation.objects.filter(id__in=data['vacation']['periods'])
            data['vacation']['periods'] = PeriodsSerializer(periods_vac, many=True).data
        if not target:
            data['target'] = data['vacation'].pop('user')
        
    if target:
        user = CustomUser.objects.get(id=target)
        data['target'] = CustomUserSerializer(user).data

    doc_for = data['target']['name']

    if added_users_ids:
        data['sub_users'] = CustomUserSerializer(CustomUser.objects.filter(id__in=added_users_ids), many=True).data

    settings = UserSettingsReadSerializer(UserSettings.objects.get(user=request.user.id)).data
        
    data['settings'] = settings
    try:
        result = generate_document_content(
            dataset=data,
            type_doc=type_doc,
            system='OGGSK',
            doc_for=doc_for
        )
        
        response = HttpResponse(
            result['content'],
            content_type=result['content_type']
        )
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quote(result["filename"])}'
        return response
        
    except Exception as e:
        return Response(
            str(e), 
            status=status.HTTP_400_BAD_REQUEST
        )
